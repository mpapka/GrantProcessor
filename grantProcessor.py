#!/usr/bin/env python3
"""
Grant Expenditure Processor

Parses COE quarterly grant expenditure files and generates per-department
summary reports with indirect cost distributions.

Usage:
    python grantProcessor.py COE-Q1AY26.xlsx -d "COMPUTER SCIENCE" -q Q1
    python grantProcessor.py COE-Q1AY26.xlsx -d all -q full-year
    python grantProcessor.py COE-Q1AY26.xlsx -d all -q Q1 --combined
    python grantProcessor.py COE-Q1AY26.xlsx  # defaults to CS, Q1
"""

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, numbers
from openpyxl.styles.colors import Color


# Quarter column mapping (1-indexed)
QUARTER_COLUMNS = {
    "Q1": 7,   # Column G: Jul-Sep
    "Q2": 8,   # Column H: Oct-Dec
    "Q3": 9,   # Column I: Jan-Mar
    "Q4": 10,  # Column J: Apr-Jun
}

# Departments to skip when running in "all" mode
SKIP_DEPARTMENTS = {"ADMINISTRATION", "COLLEGE TOTAL"}

# Department text colors (hex RGB, no leading #)
DEPT_COLORS = {
    "ERC":     "7B2D8E",  # purple
    "MIE":     "0066CC",  # blue
    "BME":     "CC0000",  # red
    "BME-COM": "E65100",  # dark orange
    "CS":      "1A7A1A",  # green
    "ChE":     "B8860B",  # dark goldenrod
    "CME":     "00838F",  # teal
    "ECE":     "6A1B9A",  # deep purple
    "AP":      "E91E63",  # pink
    "MSE":     "FF5722",  # deep orange
    "EnvE":    "009688",  # teal-green
    "CompSci": "3F51B5",  # indigo
}


def findDepartments(ws):
    """Find department section boundaries in the worksheet.

    Returns a list of (name, startRow, endRow) tuples.
    Department headers are identified as ALL-CAPS text in column B
    that isn't a fiscal year label or percentage row.
    """
    deptRows = []
    for rowNum in range(1, ws.max_row + 1):
        val = ws.cell(row=rowNum, column=2).value
        if not val or not isinstance(val, str):
            continue
        valStripped = val.strip()
        if not valStripped:
            continue
        # Department headers are ALL CAPS, at least 3 chars, and not FY/Percentage rows
        if (valStripped == valStripped.upper()
                and len(valStripped) >= 3
                and not re.match(r'^FY\s*\d', valStripped)
                and "PERCENTAGE" not in valStripped.upper()):
            deptRows.append((valStripped, rowNum))

    # Compute end rows
    departments = []
    for i, (name, start) in enumerate(deptRows):
        if i + 1 < len(deptRows):
            end = deptRows[i + 1][1] - 1
        else:
            end = ws.max_row
        departments.append((name, start, end))

    return departments


def findFySummaryRow(ws, startRow, endRow):
    """Find the FY summary row that marks the end of investigator data.

    Within a department section, the investigator data is followed by
    a summary section starting with "FY26" (or whatever current FY) in column B.
    Returns the row number of the first FY summary row, or endRow if not found.
    """
    for rowNum in range(startRow, endRow + 1):
        val = ws.cell(row=rowNum, column=2).value
        if val and isinstance(val, str):
            valStripped = val.strip()
            # Match "FY26", "FY25", etc. as summary markers
            if re.match(r'^FY\s*\d{2}$', valStripped):
                return rowNum
    return endRow + 1


def normalizeName(val):
    """Clean up investigator name by stripping trailing artifacts (backticks, etc.)."""
    if not val:
        return val
    # Strip trailing backticks, spaces, and other non-alphanumeric artifacts
    return re.sub(r'[`]+$', '', val).strip()


def isInvestigatorName(val):
    """Check if a cell value looks like an investigator name (Last, First/Last. First)."""
    if not val or not isinstance(val, str):
        return False
    val = val.strip()
    # Standard "Last, First" format
    if "," in val:
        return True
    # Handle "Gonzalo-Bello. Lander" style names (period as separator)
    # Must have a period followed by a space and a capitalized word,
    # and not be a department header (all caps) or FY label
    if re.match(r'^[A-Z][a-zA-Z-]+\.\s+[A-Z]', val):
        return True
    return False


def extractInvestigators(ws, startRow, endRow, quarter):
    """Extract per-investigator expenditure totals within a department section.

    Args:
        ws: Worksheet object
        startRow: First row of the department section
        endRow: Last row of the department section
        quarter: "Q1", "Q2", "Q3", "Q4", or "full-year"

    Returns:
        dict mapping investigator names to total expenditure amounts
    """
    # Find where the FY summary rows begin (end of investigator data)
    dataEnd = findFySummaryRow(ws, startRow, endRow)

    if quarter == "full-year":
        cols = [QUARTER_COLUMNS[q] for q in ("Q1", "Q2", "Q3", "Q4")]
    else:
        cols = [QUARTER_COLUMNS[quarter]]

    investigators = defaultdict(float)
    currentName = None

    for rowNum in range(startRow, dataEnd):
        bVal = ws.cell(row=rowNum, column=2).value

        if isInvestigatorName(bVal):
            name = normalizeName(bVal)
            currentName = name
            # Ensure this investigator exists (even if all values are 0/None)
            if name not in investigators:
                investigators[name] = 0.0
            # Accumulate expenditure from grant lines
            for col in cols:
                cellVal = ws.cell(row=rowNum, column=col).value
                if cellVal is not None and isinstance(cellVal, (int, float)):
                    investigators[name] += cellVal
        elif bVal is None:
            # Could be a subtotal row or empty row
            # Check if this is a subtotal row (has values in G but no name)
            gVal = ws.cell(row=rowNum, column=7).value
            if gVal is not None and currentName is not None:
                # This is a subtotal row - we already accumulated from
                # individual grant lines, so skip to avoid double-counting
                pass

    return dict(investigators)


def calculateDistributions(investigatorTotals, rates):
    """Calculate indirect cost distributions for each investigator.

    Args:
        investigatorTotals: dict of {name: totalExpenditure}
        rates: dict with keys 'indirect', 'campus', 'ovcr', 'coe', 'dept'

    Returns:
        List of dicts sorted by funds descending, each containing:
        name, funds, percentage, indirect, campus, ovcr, coe, dept
    """
    grandTotal = sum(investigatorTotals.values())

    results = []
    for name, funds in investigatorTotals.items():
        indirect = funds * rates["indirect"]
        pct = funds / grandTotal if grandTotal != 0 else 0
        results.append({
            "name": name,
            "funds": funds,
            "percentage": pct,
            "indirect": indirect,
            "campus": indirect * rates["campus"],
            "ovcr": indirect * rates["ovcr"],
            "coe": indirect * rates["coe"],
            "dept": indirect * rates["dept"],
        })

    results.sort(key=lambda r: r["funds"], reverse=True)
    return results


def writeOutput(results, outputPath, sheetName, deptLabel):
    """Write results to an Excel file.

    Args:
        results: List of result dicts from calculateDistributions
        outputPath: Path to the output file
        sheetName: Name for the worksheet
        deptLabel: Department abbreviation for the last column header
    """
    # Load existing workbook or create new one
    try:
        wb = openpyxl.load_workbook(outputPath)
        if sheetName in wb.sheetnames:
            del wb[sheetName]
        ws = wb.create_sheet(sheetName)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheetName

    # Remove the default empty sheet if we created a new workbook
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    headers = ["Name", "Funds", "Percentage", "Indirect (56%)",
               "Campus", "OVCR", "COE", deptLabel]

    bold = Font(bold=True)
    for colIdx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=colIdx, value=header)
        cell.font = bold

    currencyFmt = '#,##0.00'
    pctFmt = '0.00%'

    for rowIdx, r in enumerate(results, 2):
        ws.cell(row=rowIdx, column=1, value=r["name"])

        for colIdx, key in enumerate(["funds", "percentage", "indirect",
                                        "campus", "ovcr", "coe", "dept"], 2):
            cell = ws.cell(row=rowIdx, column=colIdx, value=r[key])
            if key == "percentage":
                cell.number_format = pctFmt
            else:
                cell.number_format = currencyFmt

    # Total row
    totalRow = len(results) + 2
    ws.cell(row=totalRow, column=1, value="Total").font = bold
    totalFunds = sum(r["funds"] for r in results)
    cell = ws.cell(row=totalRow, column=2, value=totalFunds)
    cell.number_format = currencyFmt
    cell.font = bold

    # Auto-size columns
    for colIdx in range(1, len(headers) + 1):
        maxLen = len(str(headers[colIdx - 1]))
        for rowIdx in range(2, totalRow + 1):
            val = ws.cell(row=rowIdx, column=colIdx).value
            if val is not None:
                maxLen = max(maxLen, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(colIdx)].width = min(maxLen + 2, 40)

    wb.save(outputPath)
    return outputPath


def writeCombinedOutput(allResults, outputPath, rates, quarterData=None):
    """Write all departments into a single sheet with a Department column.

    Args:
        allResults: List of (deptAbbr, investigatorDict) tuples
        outputPath: Path to the output file
        rates: dict with keys 'indirect', 'campus', 'ovcr', 'coe', 'dept'
        quarterData: Optional dict mapping (dept, name) to {Q1: val, ...}
    """
    hasQuarters = quarterData is not None and len(quarterData) > 0

    # Flatten all investigators with their department
    rows = []
    for deptAbbr, investigators in allResults:
        for name, funds in investigators.items():
            row = {"department": deptAbbr, "name": name, "funds": funds}
            if hasQuarters:
                qd = quarterData.get((deptAbbr, name), {})
                row["q1"] = qd.get("Q1", 0.0)
                row["q2"] = qd.get("Q2", 0.0)
                row["q3"] = qd.get("Q3", 0.0)
                row["q4"] = qd.get("Q4", 0.0)
            rows.append(row)

    # Compute college-wide total for percentage calculation
    grandTotal = sum(r["funds"] for r in rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined"

    indirectPct = int(rates["indirect"] * 100)
    headers = ["Department", "Name", "Funds", "Percentage",
               f"Indirect ({indirectPct}%)", "Campus", "OVCR", "COE", "Dept"]
    if hasQuarters:
        headers += ["Q1 Funds", "Q2 Funds", "Q3 Funds", "Q4 Funds"]

    bold = Font(bold=True)
    for colIdx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=colIdx, value=header)
        cell.font = bold

    currencyFmt = '#,##0.00'
    pctFmt = '0.00%'

    # Sort by funds descending across all departments
    rows.sort(key=lambda r: r["funds"], reverse=True)

    for rowIdx, r in enumerate(rows, 2):
        funds = r["funds"]
        indirect = funds * rates["indirect"]
        dept = r["department"]
        colorHex = DEPT_COLORS.get(dept, "000000")
        rowFont = Font(color=Color(rgb="FF" + colorHex))

        cell = ws.cell(row=rowIdx, column=1, value=dept)
        cell.font = rowFont
        cell = ws.cell(row=rowIdx, column=2, value=r["name"])
        cell.font = rowFont

        cell = ws.cell(row=rowIdx, column=3, value=funds)
        cell.number_format = currencyFmt
        cell.font = rowFont

        pct = funds / grandTotal if grandTotal != 0 else 0
        cell = ws.cell(row=rowIdx, column=4, value=pct)
        cell.number_format = pctFmt
        cell.font = rowFont

        for colIdx, val in enumerate([indirect,
                                        indirect * rates["campus"],
                                        indirect * rates["ovcr"],
                                        indirect * rates["coe"],
                                        indirect * rates["dept"]], 5):
            cell = ws.cell(row=rowIdx, column=colIdx, value=val)
            cell.number_format = currencyFmt
            cell.font = rowFont

        if hasQuarters:
            for colIdx, q in enumerate(["q1", "q2", "q3", "q4"], 10):
                cell = ws.cell(row=rowIdx, column=colIdx, value=r[q])
                cell.number_format = currencyFmt
                cell.font = rowFont

    # Total row
    totalRow = len(rows) + 2
    ws.cell(row=totalRow, column=1, value="Total").font = bold
    totalIndirect = grandTotal * rates["indirect"]
    for colIdx, val in [(3, grandTotal),
                         (5, totalIndirect),
                         (6, totalIndirect * rates["campus"]),
                         (7, totalIndirect * rates["ovcr"]),
                         (8, totalIndirect * rates["coe"]),
                         (9, totalIndirect * rates["dept"])]:
        cell = ws.cell(row=totalRow, column=colIdx, value=val)
        cell.number_format = currencyFmt
        cell.font = bold

    if hasQuarters:
        for colIdx, q in enumerate(["q1", "q2", "q3", "q4"], 10):
            val = sum(r[q] for r in rows)
            cell = ws.cell(row=totalRow, column=colIdx, value=val)
            cell.number_format = currencyFmt
            cell.font = bold

    # Auto-size columns
    for colIdx in range(1, len(headers) + 1):
        maxLen = len(str(headers[colIdx - 1]))
        for rowIdx in range(2, totalRow + 1):
            val = ws.cell(row=rowIdx, column=colIdx).value
            if val is not None:
                maxLen = max(maxLen, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(colIdx)].width = min(maxLen + 2, 40)

    wb.save(outputPath)
    return outputPath


def getDeptAbbreviation(deptName):
    """Generate a short abbreviation for a department name."""
    abbreviations = {
        "COMPUTER SCIENCE": "CS",
        "MECHANICAL AND INDUSTRIAL ENGINEERING": "MIE",
        "BIOMEDICAL ENGINEERING": "BME",
        "BME-COM": "BME-COM",
        "CHEMICAL ENGINEERING": "ChE",
        "CIVIL AND MATERIAL ENGINEERING": "CME",
        "ELECTRICAL AND COMPUTER ENGINEERING": "ECE",
        "ENERGY RESOURCE CENTER": "ERC",
        "ADMINISTRATION": "Admin",
        "APPLIED PHYSICS": "AP",
        "MATERIALS SCIENCE AND ENGINEERING": "MSE",
        "ENVIRONMENTAL ENGINEERING": "EnvE",
        "COMPUTATIONAL SCIENCES": "CompSci",
    }
    return abbreviations.get(deptName, deptName[:6])


def generateOutputFilename(deptName, quarter):
    """Generate a default output filename."""
    abbr = getDeptAbbreviation(deptName)
    if deptName == "all":
        return f"COE-Expenditures-{quarter}.xlsx"
    return f"{abbr}-Expenditures-{quarter}.xlsx"


def main():
    """CLI entry point. Parses arguments and orchestrates the processing pipeline."""
    parser = argparse.ArgumentParser(
        description="Process COE grant expenditure files into department summaries."
    )
    parser.add_argument("input_file", help="Path to the COE Excel file")
    parser.add_argument("-d", "--department", default="COMPUTER SCIENCE",
                        help='Department name or "all" (default: COMPUTER SCIENCE)')
    parser.add_argument("-q", "--quarter", default="Q1",
                        choices=["Q1", "Q2", "Q3", "Q4", "full-year"],
                        help="Quarter to report on (default: Q1)")
    parser.add_argument("-o", "--output", default=None,
                        help="Output file path (default: auto-generated)")
    parser.add_argument("--indirect-rate", type=float, default=0.56,
                        help="Indirect cost rate (default: 0.56)")
    parser.add_argument("--campus-rate", type=float, default=0.448,
                        help="Campus share of indirect (default: 0.448)")
    parser.add_argument("--ovcr-rate", type=float, default=0.07,
                        help="OVCR share of indirect (default: 0.07)")
    parser.add_argument("--coe-rate", type=float, default=0.275,
                        help="COE share of indirect (default: 0.275)")
    parser.add_argument("--dept-rate", type=float, default=0.20,
                        help="Department share of indirect (default: 0.20)")
    parser.add_argument("-c", "--combined", action="store_true",
                        help="Output all departments into a single sheet with a Department column")

    args = parser.parse_args()

    rates = {
        "indirect": args.indirect_rate,
        "campus": args.campus_rate,
        "ovcr": args.ovcr_rate,
        "coe": args.coe_rate,
        "dept": args.dept_rate,
    }

    inputPath = Path(args.input_file)
    if not inputPath.exists():
        print(f"Error: Input file not found: {inputPath}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {inputPath}...")
    wb = openpyxl.load_workbook(inputPath, data_only=True)
    ws = wb.active

    departments = findDepartments(ws)
    deptNames = [d[0] for d in departments]
    print(f"Found departments: {', '.join(deptNames)}")

    outputPath = args.output or generateOutputFilename(args.department, args.quarter)

    # Remove existing output file to start fresh
    if Path(outputPath).exists():
        Path(outputPath).unlink()

    if args.department.lower() == "all":
        targetDepts = [(n, s, e) for n, s, e in departments
                        if n not in SKIP_DEPARTMENTS]
    else:
        targetDepts = [(n, s, e) for n, s, e in departments
                        if n.upper() == args.department.upper()]
        if not targetDepts:
            print(f"Error: Department '{args.department}' not found.", file=sys.stderr)
            print(f"Available: {', '.join(deptNames)}", file=sys.stderr)
            sys.exit(1)

    if args.combined:
        # Force all departments when --combined is used
        if args.department.lower() != "all":
            targetDepts = [(n, s, e) for n, s, e in departments
                            if n not in SKIP_DEPARTMENTS]

        combinedOutput = args.output or f"COE-Combined-{args.quarter}.xlsx"
        if Path(combinedOutput).exists():
            Path(combinedOutput).unlink()

        allResults = []
        # When processing full-year, also extract per-quarter breakdowns
        quarterData = {} if args.quarter == "full-year" else None
        for deptName, startRow, endRow in targetDepts:
            abbr = getDeptAbbreviation(deptName)
            print(f"Processing {deptName} ({abbr})...")
            investigators = extractInvestigators(ws, startRow, endRow, args.quarter)
            allResults.append((abbr, investigators))

            if quarterData is not None:
                for q in ("Q1", "Q2", "Q3", "Q4"):
                    qInv = extractInvestigators(ws, startRow, endRow, q)
                    for name, funds in qInv.items():
                        qKey = (abbr, name)
                        if qKey not in quarterData:
                            quarterData[qKey] = {}
                        quarterData[qKey][q] = funds

            total = sum(investigators.values())
            active = sum(1 for v in investigators.values() if v != 0)
            print(f"  {len(investigators)} investigators ({active} active), "
                  f"total expenditures: ${total:,.2f}")

        writeCombinedOutput(allResults, combinedOutput, rates, quarterData)
        print(f"Output saved to: {combinedOutput}")
    else:
        for deptName, startRow, endRow in targetDepts:
            abbr = getDeptAbbreviation(deptName)
            print(f"Processing {deptName} ({abbr})...")

            investigators = extractInvestigators(ws, startRow, endRow, args.quarter)
            results = calculateDistributions(investigators, rates)

            sheetName = abbr if args.department.lower() == "all" else "Sheet1"
            writeOutput(results, outputPath, sheetName, abbr)

            total = sum(r["funds"] for r in results)
            active = sum(1 for r in results if r["funds"] != 0)
            print(f"  {len(results)} investigators ({active} active), "
                  f"total expenditures: ${total:,.2f}")

        print(f"Output saved to: {outputPath}")


if __name__ == "__main__":
    main()
