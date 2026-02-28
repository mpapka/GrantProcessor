#!/usr/bin/env python3
"""
Grant Summary Plot Generator

Reads the combined Excel output from grantProcessor.py and generates
presentation-ready summary plots in PNG, PDF, and interactive HTML.

Usage:
    python plotGrantSummary.py COE-Combined-Q1.xlsx
    python plotGrantSummary.py COE-Combined-Q1.xlsx --top 10 --format png
    python plotGrantSummary.py COE-Combined-Q1.xlsx --format html -o plots/
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import plotly.graph_objects as go

from grantProcessor import DEPT_COLORS


def hexToRgb(hexStr):
    """Convert a 6-char hex string (no #) to an (r, g, b) tuple scaled 0-1."""
    r = int(hexStr[0:2], 16) / 255
    g = int(hexStr[2:4], 16) / 255
    b = int(hexStr[4:6], 16) / 255
    return (r, g, b)


def hexToPlotly(hexStr):
    """Convert a 6-char hex string to a plotly-compatible '#RRGGBB' string."""
    return f"#{hexStr}"


def deptColor(dept, forPlotly=False):
    """Return the color for a department abbreviation."""
    hexStr = DEPT_COLORS.get(dept, "555555")
    if forPlotly:
        return hexToPlotly(hexStr)
    return hexToRgb(hexStr)


def readCombinedData(filePath):
    """Read the Combined sheet from the output Excel file.

    Returns a list of dicts with keys: department, name, funds, percentage,
    indirect, campus, ovcr, coe, dept. If per-quarter columns exist
    (Q1 Funds .. Q4 Funds in columns 10-13), also includes q1-q4.
    The Total row is excluded.
    """
    wb = openpyxl.load_workbook(filePath, data_only=True)
    ws = wb["Combined"]

    # Check if per-quarter columns exist
    hasQuarters = (ws.cell(row=1, column=10).value or "").startswith("Q1")

    rows = []
    for rowNum in range(2, ws.max_row + 1):
        name = ws.cell(row=rowNum, column=2).value
        if name is None or (isinstance(name, str) and name.strip() == "Total"):
            continue
        department = ws.cell(row=rowNum, column=1).value
        if department is None:
            continue

        funds = ws.cell(row=rowNum, column=3).value or 0
        percentage = ws.cell(row=rowNum, column=4).value or 0
        indirect = ws.cell(row=rowNum, column=5).value or 0
        campus = ws.cell(row=rowNum, column=6).value or 0
        ovcr = ws.cell(row=rowNum, column=7).value or 0
        coe = ws.cell(row=rowNum, column=8).value or 0
        dept = ws.cell(row=rowNum, column=9).value or 0

        row = {
            "department": str(department).strip(),
            "name": str(name).strip(),
            "funds": float(funds),
            "percentage": float(percentage),
            "indirect": float(indirect),
            "campus": float(campus),
            "ovcr": float(ovcr),
            "coe": float(coe),
            "dept": float(dept),
        }

        if hasQuarters:
            row["q1"] = float(ws.cell(row=rowNum, column=10).value or 0)
            row["q2"] = float(ws.cell(row=rowNum, column=11).value or 0)
            row["q3"] = float(ws.cell(row=rowNum, column=12).value or 0)
            row["q4"] = float(ws.cell(row=rowNum, column=13).value or 0)

        rows.append(row)

    wb.close()
    return rows


def formatDollars(value):
    """Format a number as a dollar string."""
    if abs(value) >= 1_000_000:
        return f"${value / 1_000_000:,.1f}M"
    if abs(value) >= 1_000:
        return f"${value / 1_000:,.0f}K"
    return f"${value:,.0f}"


def plotTopInvestigators(data, topN=20):
    """Create top-N investigators horizontal bar chart.

    Returns (matplotlibFig, plotlyFig).
    """
    sorted_data = sorted(data, key=lambda r: r["funds"], reverse=True)[:topN]
    # Reverse for horizontal bar chart (top item at top)
    sorted_data = sorted_data[::-1]

    names = [f"{r['name']} ({r['department']})" for r in sorted_data]
    funds = [r["funds"] for r in sorted_data]
    colors = [deptColor(r["department"]) for r in sorted_data]
    plotlyColors = [deptColor(r["department"], forPlotly=True) for r in sorted_data]

    # --- Matplotlib ---
    figHeight = max(6, topN * 0.35)
    fig, ax = plt.subplots(figsize=(12, figHeight))
    bars = ax.barh(range(len(names)), funds, color=colors)
    ax.set_yticks(range(len(names)))
    ax.set_yticklabels(names, fontsize=9)
    ax.set_xlabel("Expenditures ($)")
    ax.set_title(f"Top {topN} Investigators by Expenditures")
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: formatDollars(x)))
    ax.invert_yaxis()
    fig.tight_layout()

    # --- Plotly ---
    pFig = go.Figure(go.Bar(
        y=names[::-1],
        x=funds[::-1],
        orientation="h",
        marker_color=plotlyColors[::-1],
        text=[formatDollars(f) for f in funds[::-1]],
        textposition="outside",
    ))
    pFig.update_layout(
        title=f"Top {topN} Investigators by Expenditures",
        xaxis_title="Expenditures ($)",
        height=max(500, topN * 30),
        margin=dict(l=250),
        yaxis=dict(autorange="reversed"),
    )

    return fig, pFig


def plotDeptTotals(data):
    """Create department totals vertical bar chart.

    Returns (matplotlibFig, plotlyFig).
    """
    # Aggregate by department
    deptTotals = {}
    for r in data:
        dept = r["department"]
        deptTotals[dept] = deptTotals.get(dept, 0) + r["funds"]

    # Sort descending
    sorted_depts = sorted(deptTotals.items(), key=lambda x: x[1], reverse=True)
    depts = [d[0] for d in sorted_depts]
    totals = [d[1] for d in sorted_depts]
    colors = [deptColor(d) for d in depts]
    plotlyColors = [deptColor(d, forPlotly=True) for d in depts]

    # --- Matplotlib ---
    fig, ax = plt.subplots(figsize=(10, 6))
    bars = ax.bar(range(len(depts)), totals, color=colors)
    ax.set_xticks(range(len(depts)))
    ax.set_xticklabels(depts, rotation=45, ha="right")
    ax.set_ylabel("Expenditures ($)")
    ax.set_title("Total Expenditures by Department")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: formatDollars(x)))

    # Dollar labels on top of bars
    for bar, total in zip(bars, totals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(),
                formatDollars(total), ha="center", va="bottom", fontsize=9)

    fig.tight_layout()

    # --- Plotly ---
    pFig = go.Figure(go.Bar(
        x=depts,
        y=totals,
        marker_color=plotlyColors,
        text=[formatDollars(t) for t in totals],
        textposition="outside",
    ))
    pFig.update_layout(
        title="Total Expenditures by Department",
        yaxis_title="Expenditures ($)",
    )

    return fig, pFig


def plotIndirectBreakdown(data):
    """Create stacked bar chart of indirect cost distribution by department.

    Returns (matplotlibFig, plotlyFig).
    """
    # Aggregate by department
    deptIndirect = {}
    for r in data:
        dept = r["department"]
        if dept not in deptIndirect:
            deptIndirect[dept] = {"campus": 0, "ovcr": 0, "coe": 0, "dept": 0}
        deptIndirect[dept]["campus"] += r["campus"]
        deptIndirect[dept]["ovcr"] += r["ovcr"]
        deptIndirect[dept]["coe"] += r["coe"]
        deptIndirect[dept]["dept"] += r["dept"]

    # Sort by total indirect descending
    sorted_depts = sorted(deptIndirect.items(),
                          key=lambda x: sum(x[1].values()), reverse=True)
    depts = [d[0] for d in sorted_depts]
    campusVals = [d[1]["campus"] for d in sorted_depts]
    ovcrVals = [d[1]["ovcr"] for d in sorted_depts]
    coeVals = [d[1]["coe"] for d in sorted_depts]
    deptVals = [d[1]["dept"] for d in sorted_depts]

    categoryColors = {
        "Campus": "#2196F3",
        "OVCR": "#FF9800",
        "COE": "#4CAF50",
        "Dept": "#9C27B0",
    }

    # --- Matplotlib ---
    fig, ax = plt.subplots(figsize=(10, 6))
    x = range(len(depts))
    bottom = [0] * len(depts)

    for label, vals in [("Campus", campusVals), ("OVCR", ovcrVals),
                        ("COE", coeVals), ("Dept", deptVals)]:
        ax.bar(x, vals, bottom=bottom, label=label, color=categoryColors[label])
        bottom = [b + v for b, v in zip(bottom, vals)]

    ax.set_xticks(x)
    ax.set_xticklabels(depts, rotation=45, ha="right")
    ax.set_ylabel("Indirect Cost ($)")
    ax.set_title("Indirect Cost Distribution by Department")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: formatDollars(x)))
    ax.legend()
    fig.tight_layout()

    # --- Plotly ---
    pFig = go.Figure()
    for label, vals in [("Campus", campusVals), ("OVCR", ovcrVals),
                        ("COE", coeVals), ("Dept", deptVals)]:
        pFig.add_trace(go.Bar(
            x=depts,
            y=vals,
            name=label,
            marker_color=categoryColors[label],
            text=[formatDollars(v) for v in vals],
            textposition="inside",
        ))
    pFig.update_layout(
        barmode="stack",
        title="Indirect Cost Distribution by Department",
        yaxis_title="Indirect Cost ($)",
    )

    return fig, pFig


def writeDashboard(data, outputDir):
    """Generate a self-contained dashboard.html with tabbed D3.js visualizations.

    Produces a single HTML file embedding investigator data as JSON and using
    D3.js v7 (loaded via CDN) for three interactive SVG charts:
      Tab 1 — Top Investigators (horizontal bar, animated enter/update/exit)
      Tab 2 — Department Totals (stacked per-investigator vertical bars)
      Tab 3 — Indirect Cost Breakdown (stacked category bars)

    Controls: top-N selector, sort toggle, department filter checkboxes,
    and an optional quarter selector when per-quarter data is available.
    """
    outputDir = Path(outputDir)
    outputDir.mkdir(parents=True, exist_ok=True)

    # Serialize data and colors for embedding
    hasQuarters = len(data) > 0 and "q1" in data[0]
    dataJson = json.dumps(data)
    colorsJson = json.dumps({k: f"#{v}" for k, v in DEPT_COLORS.items()})

    categoryColors = {
        "Campus": "#2196F3",
        "OVCR": "#FF9800",
        "COE": "#4CAF50",
        "Dept": "#9C27B0",
    }
    categoryColorsJson = json.dumps(categoryColors)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Grant Expenditure Dashboard</title>
<script src="https://d3js.org/d3.v7.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
         background: #f5f7fa; color: #333; }}
  .header {{ background: #1a237e; color: #fff; padding: 20px 32px; }}
  .header h1 {{ font-size: 22px; font-weight: 600; }}
  .tab-bar {{ display: flex; background: #fff; border-bottom: 2px solid #e0e0e0;
              padding: 0 24px; }}
  .tab {{ padding: 14px 24px; cursor: pointer; font-size: 14px; font-weight: 500;
          color: #666; border-bottom: 3px solid transparent; transition: all 0.2s;
          user-select: none; }}
  .tab:hover {{ color: #1a237e; background: #f0f0ff; }}
  .tab.active {{ color: #1a237e; border-bottom-color: #1a237e; }}
  .tab-content {{ display: none; padding: 24px 32px; }}
  .tab-content.active {{ display: block; }}
  .controls {{ display: flex; gap: 20px; align-items: flex-start; flex-wrap: wrap;
               margin-bottom: 16px; padding: 16px; background: #fff;
               border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
  .control-group {{ display: flex; flex-direction: column; gap: 4px; }}
  .control-group label {{ font-size: 12px; font-weight: 600; color: #666;
                          text-transform: uppercase; letter-spacing: 0.5px; }}
  .control-group select {{ padding: 6px 12px; border: 1px solid #ccc;
                           border-radius: 4px; font-size: 14px; }}
  .dept-filters {{ display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }}
  .dept-filters label.filter-label {{ font-size: 12px; font-weight: 600; color: #666;
                                       text-transform: uppercase; letter-spacing: 0.5px;
                                       margin-right: 4px; }}
  .dept-check {{ display: flex; align-items: center; gap: 4px; font-size: 13px;
                 cursor: pointer; }}
  .dept-check input {{ cursor: pointer; }}
  .chart-box {{ background: #fff; border-radius: 8px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.08); padding: 16px;
                overflow: hidden; }}
  .chart-box svg {{ display: block; }}
  .chart-title {{ font-size: 16px; font-weight: 600; color: #333;
                  text-anchor: middle; }}
  .axis text {{ font-size: 12px; fill: #555; }}
  .axis path, .axis line {{ stroke: #ccc; }}
  .bar-label {{ font-size: 11px; fill: #333; }}
  .tooltip {{ position: fixed; pointer-events: none; background: #fff;
              border: 1px solid #ddd; border-radius: 6px; padding: 10px 14px;
              box-shadow: 0 4px 12px rgba(0,0,0,0.12); font-size: 13px;
              line-height: 1.5; opacity: 0; transition: opacity 0.15s;
              z-index: 1000; max-width: 300px; }}
  .tooltip .tt-name {{ font-weight: 600; color: #1a237e; }}
  .tooltip .tt-dept {{ color: #666; font-size: 12px; }}
  .tooltip .tt-amount {{ font-size: 15px; font-weight: 600; }}
  .legend {{ display: flex; gap: 16px; margin-top: 8px; justify-content: center; }}
  .legend-item {{ display: flex; align-items: center; gap: 5px; font-size: 13px; }}
  .legend-swatch {{ width: 14px; height: 14px; border-radius: 2px; }}
</style>
</head>
<body>

<div class="header">
  <h1>Grant Expenditure Dashboard</h1>
</div>

<div class="tab-bar">
  <div class="tab active" data-tab="topInv">Top Investigators</div>
  <div class="tab" data-tab="deptTot">Department Totals</div>
  <div class="tab" data-tab="indirect">Indirect Breakdown</div>
</div>

{"" if not hasQuarters else '''<div class="controls" style="margin: 12px 32px 0; padding: 10px 16px;">
  <div class="control-group">
    <label>Quarter</label>
    <select id="quarterSelect">
      <option value="all" selected>Full Year</option>
      <option value="q1">Q1 (Jul\u2013Sep)</option>
      <option value="q2">Q2 (Oct\u2013Dec)</option>
      <option value="q3">Q3 (Jan\u2013Mar)</option>
      <option value="q4">Q4 (Apr\u2013Jun)</option>
    </select>
  </div>
</div>
'''}
<!-- Tab 1: Top Investigators -->
<div id="topInv" class="tab-content active">
  <div class="controls">
    <div class="control-group">
      <label>Show Top</label>
      <select id="topNSelect">
        <option value="5">5</option>
        <option value="10">10</option>
        <option value="15">15</option>
        <option value="20" selected>20</option>
        <option value="30">30</option>
        <option value="0">All</option>
      </select>
    </div>
    <div class="control-group">
      <label>Sort By</label>
      <select id="sortSelect">
        <option value="funds">By Funds</option>
        <option value="department">By Department</option>
      </select>
    </div>
    <div class="control-group">
      <label class="filter-label">Departments</label>
      <div class="dept-filters" id="deptFilters"></div>
    </div>
  </div>
  <div class="chart-box" id="chartTopInv"></div>
</div>

<!-- Tab 2: Department Totals -->
<div id="deptTot" class="tab-content">
  <div class="chart-box" id="chartDeptTot"></div>
</div>

<!-- Tab 3: Indirect Breakdown -->
<div id="indirect" class="tab-content">
  <div class="chart-box" id="chartIndirect"></div>
  <div class="legend" id="indirectLegend"></div>
</div>

<!-- Shared tooltip -->
<div class="tooltip" id="tooltip"></div>

<script>
const DATA = {dataJson};
const DEPT_COLORS = {colorsJson};
const CAT_COLORS = {categoryColorsJson};
const CAT_KEYS = ["campus", "ovcr", "coe", "dept"];
const CAT_LABELS = {{ campus: "Campus", ovcr: "OVCR", coe: "COE", dept: "Dept" }};
const DEFAULT_COLOR = "#555555";

function getDeptColor(dept) {{
  return DEPT_COLORS[dept] || DEFAULT_COLOR;
}}

function formatDollars(v) {{
  const abs = Math.abs(v);
  if (abs >= 1e6) return "$" + (v / 1e6).toFixed(1) + "M";
  if (abs >= 1e3) return "$" + Math.round(v / 1e3).toLocaleString() + "K";
  return "$" + Math.round(v).toLocaleString();
}}

/* ---------- Quarter-aware data ---------- */
const HAS_QUARTERS = DATA.length > 0 && "q1" in DATA[0];
function getSelectedQuarter() {{
  const el = document.getElementById("quarterSelect");
  return el ? el.value : "all";
}}
function getFunds(d) {{
  const q = getSelectedQuarter();
  if (q === "all") return d.funds;
  return d[q] || 0;
}}
// Build a quarter-filtered view of DATA with recalculated indirect splits
function getActiveData() {{
  const q = getSelectedQuarter();
  if (q === "all") return DATA;
  return DATA.map(d => {{
    const f = d[q] || 0;
    const ratio = d.funds !== 0 ? f / d.funds : 0;
    return {{
      ...d,
      funds: f,
      indirect: d.indirect * ratio,
      campus: d.campus * ratio,
      ovcr: d.ovcr * ratio,
      coe: d.coe * ratio,
      dept: d.dept * ratio,
    }};
  }});
}}

/* ---------- Tooltip helpers ---------- */
const tooltip = d3.select("#tooltip");
function showTooltip(evt, html) {{
  tooltip.html(html).style("opacity", 1);
  positionTooltip(evt);
}}
function positionTooltip(evt) {{
  const tt = document.getElementById("tooltip");
  const ttRect = tt.getBoundingClientRect();
  let x = evt.clientX + 14;
  let y = evt.clientY - 10;
  if (x + ttRect.width > window.innerWidth - 8) x = evt.clientX - ttRect.width - 14;
  if (y + ttRect.height > window.innerHeight - 8) y = window.innerHeight - ttRect.height - 8;
  if (y < 8) y = 8;
  tooltip.style("left", x + "px").style("top", y + "px");
}}
function hideTooltip() {{
  tooltip.style("opacity", 0);
}}

/* ---------- Tab switching ---------- */
document.querySelectorAll(".tab").forEach(tab => {{
  tab.addEventListener("click", () => {{
    document.querySelectorAll(".tab").forEach(t => t.classList.remove("active"));
    document.querySelectorAll(".tab-content").forEach(tc => tc.classList.remove("active"));
    tab.classList.add("active");
    document.getElementById(tab.dataset.tab).classList.add("active");
    scheduleResize();
  }});
}});

/* ---------- Department filter checkboxes ---------- */
const allDepts = [...new Set(DATA.map(d => d.department))].sort();
const filtersDiv = document.getElementById("deptFilters");
allDepts.forEach(dept => {{
  const lbl = document.createElement("label");
  lbl.className = "dept-check";
  const cb = document.createElement("input");
  cb.type = "checkbox";
  cb.checked = true;
  cb.value = dept;
  cb.addEventListener("change", renderTopInvestigators);
  lbl.appendChild(cb);
  lbl.appendChild(document.createTextNode(dept));
  filtersDiv.appendChild(lbl);
}});

/* ---------- Resize handling ---------- */
let resizeTimer;
function scheduleResize() {{
  clearTimeout(resizeTimer);
  resizeTimer = setTimeout(() => {{
    // Reset persistent SVG on resize so it adapts to new container width
    if (topInvState.svg) {{ topInvState.svg.remove(); topInvState.svg = null; }}
    renderTopInvestigators(true);
    renderDeptTotals();
    renderIndirectBreakdown();
  }}, 120);
}}
new ResizeObserver(scheduleResize).observe(document.body);

/* ---------- Tab 1: Top Investigators (horizontal bar chart) ---------- */
// Persistent SVG — created once, updated via enter/update/exit
let topInvState = {{ svg: null, g: null, xAxisG: null, titleEl: null }};

function renderTopInvestigators(instant) {{
  const container = document.getElementById("chartTopInv");
  const topN = parseInt(document.getElementById("topNSelect").value);
  const sortBy = document.getElementById("sortSelect").value;
  const dur = (instant === true) ? 0 : 800;

  const activeDepts = new Set();
  filtersDiv.querySelectorAll("input:checked").forEach(cb => activeDepts.add(cb.value));

  let filtered = getActiveData().filter(d => activeDepts.has(d.department));

  if (sortBy === "funds") {{
    filtered.sort((a, b) => b.funds - a.funds);
  }} else {{
    filtered.sort((a, b) => {{
      if (a.department < b.department) return -1;
      if (a.department > b.department) return 1;
      return b.funds - a.funds;
    }});
  }}

  if (topN > 0) filtered = filtered.slice(0, topN);

  // Unique key per investigator
  const key = d => d.name + "|" + d.department;

  const count = filtered.length;
  const margin = {{ top: 40, right: 100, bottom: 40, left: 220 }};
  const width = Math.max(200, container.clientWidth - margin.left - margin.right);
  const barHeight = 26;
  const barPad = 4;
  const height = Math.max(300, count * (barHeight + barPad));
  const fullW = width + margin.left + margin.right;
  const fullH = height + margin.top + margin.bottom;

  const maxFunds = d3.max(filtered, d => d.funds) || 1;

  const x = d3.scaleLinear()
    .domain([0, maxFunds * 1.12])
    .range([0, width]);

  const y = d3.scaleBand()
    .domain(filtered.map(key))
    .range([0, height])
    .padding(0.15);

  // Create persistent SVG on first call
  if (!topInvState.svg) {{
    topInvState.svg = d3.select(container).append("svg");
    topInvState.g = topInvState.svg.append("g")
      .attr("transform", `translate(${{margin.left}},${{margin.top}})`);
    topInvState.xAxisG = topInvState.g.append("g").attr("class", "axis");
    topInvState.titleEl = topInvState.svg.append("text").attr("class", "chart-title");
  }}

  const svg = topInvState.svg;
  const g = topInvState.g;

  svg.transition().duration(dur)
    .attr("width", fullW)
    .attr("height", fullH);

  // Title
  topInvState.titleEl
    .attr("x", fullW / 2).attr("y", 26)
    .text("Top " + count + " Investigators by Expenditures");

  // X axis
  topInvState.xAxisG
    .attr("transform", `translate(0,${{height}})`)
    .transition().duration(dur)
    .call(d3.axisBottom(x).ticks(6).tickFormat(formatDollars));

  // --- BARS: enter / update / exit ---
  const bars = g.selectAll(".bar")
    .data(filtered, key);

  // EXIT — slide right and fade out, then remove
  bars.exit()
    .transition().duration(dur * 0.6).ease(d3.easeCubicIn)
    .attr("x", width + 20)
    .style("opacity", 0)
    .remove();

  // ENTER — new bars start at width 0
  const barsEnter = bars.enter().append("rect")
    .attr("class", "bar")
    .attr("x", 0)
    .attr("width", 0)
    .attr("height", y.bandwidth())
    .attr("y", d => y(key(d)))
    .attr("fill", d => getDeptColor(d.department))
    .attr("rx", 3)
    .style("opacity", 0)
    .on("mouseover", function(evt, d) {{
      d3.select(this).style("opacity", 0.8);
      showTooltip(evt,
        `<div class="tt-name">${{d.name}}</div>` +
        `<div class="tt-dept">${{d.department}}</div>` +
        `<div class="tt-amount">${{formatDollars(d.funds)}}</div>`);
    }})
    .on("mousemove", positionTooltip)
    .on("mouseout", function() {{
      d3.select(this).style("opacity", 1);
      hideTooltip();
    }});

  // ENTER + UPDATE — slide to new position and width
  barsEnter.merge(bars)
    .transition().duration(dur).delay((d, i) => i * 15).ease(d3.easeCubicOut)
    .attr("y", d => y(key(d)))
    .attr("height", y.bandwidth())
    .attr("width", d => x(d.funds))
    .style("opacity", 1);

  // --- LABELS: enter / update / exit ---
  const labels = g.selectAll(".y-label")
    .data(filtered, key);

  labels.exit()
    .transition().duration(dur * 0.5)
    .style("opacity", 0)
    .remove();

  const labelsEnter = labels.enter().append("text")
    .attr("class", "y-label")
    .attr("x", -8)
    .attr("dy", "0.35em")
    .attr("text-anchor", "end")
    .style("font-size", "11px")
    .style("fill", "#555")
    .style("opacity", 0);

  labelsEnter.merge(labels)
    .text(d => d.name + " (" + d.department + ")")
    .transition().duration(dur).delay((d, i) => i * 15).ease(d3.easeCubicOut)
    .attr("y", d => y(key(d)) + y.bandwidth() / 2)
    .style("opacity", 1);

  // --- DOLLAR LABELS: enter / update / exit ---
  const dlabels = g.selectAll(".bar-label")
    .data(filtered, key);

  dlabels.exit()
    .transition().duration(dur * 0.5)
    .style("opacity", 0)
    .remove();

  const dlabelsEnter = dlabels.enter().append("text")
    .attr("class", "bar-label")
    .attr("dy", "0.35em")
    .style("opacity", 0);

  dlabelsEnter.merge(dlabels)
    .text(d => formatDollars(d.funds))
    .transition().duration(dur).delay((d, i) => i * 15).ease(d3.easeCubicOut)
    .attr("x", d => x(d.funds) + 6)
    .attr("y", d => y(key(d)) + y.bandwidth() / 2)
    .style("opacity", 1);
}}

document.getElementById("topNSelect").addEventListener("change", renderTopInvestigators);
document.getElementById("sortSelect").addEventListener("change", renderTopInvestigators);

/* ---------- Tab 2: Department Totals (stacked individuals) ---------- */
function lightenColor(hex, amount) {{
  // Blend hex color toward white by amount (0 = original, 1 = white)
  const r = parseInt(hex.slice(1,3), 16);
  const g = parseInt(hex.slice(3,5), 16);
  const b = parseInt(hex.slice(5,7), 16);
  const lr = Math.round(r + (255 - r) * amount);
  const lg = Math.round(g + (255 - g) * amount);
  const lb = Math.round(b + (255 - b) * amount);
  return `rgb(${{lr}},${{lg}},${{lb}})`;
}}

function renderDeptTotals() {{
  const container = document.getElementById("chartDeptTot");

  // Group investigators by department
  const activeData = getActiveData();
  const grouped = {{}};
  activeData.forEach(d => {{
    if (!grouped[d.department]) grouped[d.department] = [];
    grouped[d.department].push(d);
  }});

  // Sort departments by net total descending
  const deptTotals = Object.entries(grouped).map(([dept, inv]) => ({{
    dept, investigators: inv, total: inv.reduce((s, d) => s + d.funds, 0)
  }}));
  deptTotals.sort((a, b) => b.total - a.total);

  // Build stacked segments — positives stack up, negatives stack down from peak
  const segments = [];
  let globalPeak = 0;
  deptTotals.forEach(dt => {{
    const positives = dt.investigators.filter(d => d.funds > 0).sort((a, b) => b.funds - a.funds);
    const negatives = dt.investigators.filter(d => d.funds < 0).sort((a, b) => a.funds - b.funds);
    const posCount = positives.length;
    const negCount = negatives.length;

    // Stack positives from 0 upward
    let cum = 0;
    positives.forEach((inv, i) => {{
      const y0 = cum;
      const y1 = cum + inv.funds;
      const lightness = posCount > 1 ? (i / (posCount - 1)) * 0.55 : 0;
      segments.push({{
        dept: dt.dept, name: inv.name, funds: inv.funds,
        y0, y1, rank: i + 1, total: dt.investigators.length,
        color: lightenColor(getDeptColor(dt.dept), lightness),
        deptTotal: dt.total, isNeg: false
      }});
      cum = y1;
    }});
    dt.grossPositive = cum;
    if (cum > globalPeak) globalPeak = cum;

    // Stack negatives downward from grossPositive
    negatives.forEach((inv, i) => {{
      const y0 = cum;
      const y1 = cum + inv.funds; // funds < 0, so y1 < y0
      const lightness = negCount > 1 ? (i / (negCount - 1)) * 0.35 : 0;
      segments.push({{
        dept: dt.dept, name: inv.name, funds: inv.funds,
        y0: y1, y1: y0, // swap so y0 < y1 (bottom < top) for rendering
        rank: posCount + i + 1, total: dt.investigators.length,
        color: lightenColor("#ef5350", lightness),
        deptTotal: dt.total, isNeg: true
      }});
      cum = y1;
    }});
  }});

  const depts = deptTotals.map(d => d.dept);
  const maxVal = Math.max(d3.max(deptTotals, d => d.total), globalPeak) || 1;

  const margin = {{ top: 50, right: 20, bottom: 120, left: 80 }};
  const width = container.clientWidth - margin.left - margin.right;
  const height = 400;

  d3.select(container).select("svg").remove();

  const svg = d3.select(container).append("svg")
    .attr("width", width + margin.left + margin.right)
    .attr("height", height + margin.top + margin.bottom);

  const g = svg.append("g")
    .attr("transform", `translate(${{margin.left}},${{margin.top}})`);

  svg.append("text")
    .attr("class", "chart-title")
    .attr("x", (width + margin.left + margin.right) / 2)
    .attr("y", 24)
    .text("Total Expenditures by Department");

  const x = d3.scaleBand()
    .domain(depts)
    .range([0, width])
    .padding(0.25);

  const y = d3.scaleLinear()
    .domain([0, maxVal * 1.18])
    .range([height, 0]);

  // X axis
  g.append("g")
    .attr("class", "axis")
    .attr("transform", `translate(0,${{height}})`)
    .call(d3.axisBottom(x))
    .selectAll("text")
    .attr("transform", "rotate(-40)")
    .style("text-anchor", "end");

  // Y axis
  g.append("g")
    .attr("class", "axis")
    .call(d3.axisLeft(y).ticks(6).tickFormat(formatDollars));

  // Y axis label
  g.append("text")
    .attr("class", "axis")
    .attr("transform", "rotate(-90)")
    .attr("y", -65)
    .attr("x", -height / 2)
    .attr("text-anchor", "middle")
    .style("font-size", "13px")
    .text("Expenditures ($)");

  // Stats below x-axis labels
  const statsY = height + 62;
  const statLines = [
    {{ label: "PIs", fn: d => d.investigators.length }},
    {{ label: "Avg", fn: d => formatDollars(d.total / d.investigators.length) }},
    {{ label: "Med", fn: d => {{
      const sorted = d.investigators.map(i => i.funds).sort((a, b) => a - b);
      const mid = Math.floor(sorted.length / 2);
      return formatDollars(sorted.length % 2 ? sorted[mid] : (sorted[mid - 1] + sorted[mid]) / 2);
    }} }},
  ];
  statLines.forEach((stat, si) => {{
    // Row label on the left
    g.append("text")
      .attr("x", -8)
      .attr("y", statsY + si * 16)
      .attr("text-anchor", "end")
      .style("font-size", "11px")
      .style("font-weight", "600")
      .style("fill", "#888")
      .text(stat.label);
    // Value per department
    g.selectAll(".stat-" + si)
      .data(deptTotals)
      .enter().append("text")
      .attr("class", "stat-" + si)
      .attr("x", d => x(d.dept) + x.bandwidth() / 2)
      .attr("y", statsY + si * 16)
      .attr("text-anchor", "middle")
      .style("font-size", "11px")
      .style("fill", "#555")
      .text(d => stat.fn(d));
  }});

  // Stacked investigator segments
  g.selectAll(".seg")
    .data(segments)
    .enter().append("rect")
    .attr("class", "seg")
    .attr("x", d => x(d.dept))
    .attr("width", x.bandwidth())
    .attr("y", height)
    .attr("height", 0)
    .attr("fill", d => d.color)
    .attr("stroke", "#fff")
    .attr("stroke-width", 0.5)
    .on("mouseover", function(evt, d) {{
      d3.select(this).style("opacity", 0.8).attr("stroke", "#333").attr("stroke-width", 1.5);
      const negLabel = d.isNeg ? ' <span style="color:#ef5350">(correction)</span>' : '';
      showTooltip(evt,
        `<div class="tt-name">${{d.name}}${{negLabel}}</div>` +
        `<div class="tt-dept">${{d.dept}} \u2014 #${{d.rank}} of ${{d.total}}</div>` +
        `<div class="tt-amount">${{formatDollars(d.funds)}}</div>` +
        `<div style="font-size:12px;color:#888;margin-top:4px;border-top:1px solid #eee;padding-top:4px">Dept net total: ${{formatDollars(d.deptTotal)}}</div>`);
    }})
    .on("mousemove", positionTooltip)
    .on("mouseout", function() {{
      d3.select(this).style("opacity", 1).attr("stroke", "#fff").attr("stroke-width", 0.5);
      hideTooltip();
    }})
    .transition().duration(500).ease(d3.easeCubicOut)
    .attr("y", d => y(d.y1))
    .attr("height", d => y(d.y0) - y(d.y1));

  // For departments with corrections: horizontal line at net total
  deptTotals.filter(d => d.grossPositive !== d.total).forEach(d => {{
    g.append("line")
      .attr("x1", x(d.dept) - 4)
      .attr("x2", x(d.dept) + x.bandwidth() + 4)
      .attr("y1", y(d.total))
      .attr("y2", y(d.total))
      .attr("stroke", "#333")
      .attr("stroke-width", 2)
      .attr("stroke-dasharray", "6,3");
  }});

  // Dollar labels directly above each bar with white outline for readability
  // For depts with corrections, label goes above the gross positive peak
  function labelTop(d) {{
    return d.grossPositive > d.total ? y(d.grossPositive) - 8 : y(d.total) - 8;
  }}

  g.selectAll(".bar-label-bg")
    .data(deptTotals)
    .enter().append("text")
    .attr("class", "bar-label-bg")
    .attr("x", d => x(d.dept) + x.bandwidth() / 2)
    .attr("y", labelTop)
    .attr("text-anchor", "middle")
    .text(d => formatDollars(d.total) + (d.grossPositive > d.total ? " net" : ""))
    .style("font-size", "11px")
    .style("font-weight", "600")
    .style("stroke", "#fff")
    .style("stroke-width", 4)
    .style("stroke-linejoin", "round")
    .style("paint-order", "stroke")
    .style("opacity", 0)
    .transition().delay(300).duration(300)
    .style("opacity", 1);

  g.selectAll(".bar-label")
    .data(deptTotals)
    .enter().append("text")
    .attr("class", "bar-label")
    .attr("x", d => x(d.dept) + x.bandwidth() / 2)
    .attr("y", labelTop)
    .attr("text-anchor", "middle")
    .text(d => formatDollars(d.total) + (d.grossPositive > d.total ? " net" : ""))
    .style("font-size", "11px")
    .style("font-weight", "600")
    .style("opacity", 0)
    .transition().delay(300).duration(300)
    .style("opacity", 1);
}}

/* ---------- Tab 3: Indirect Breakdown (stacked bar chart) ---------- */
function renderIndirectBreakdown() {{
  const container = document.getElementById("chartIndirect");
  const activeData = getActiveData();
  const agg = {{}};
  activeData.forEach(d => {{
    if (!agg[d.department]) agg[d.department] = {{ campus: 0, ovcr: 0, coe: 0, dept: 0 }};
    agg[d.department].campus += d.campus;
    agg[d.department].ovcr += d.ovcr;
    agg[d.department].coe += d.coe;
    agg[d.department].dept += d.dept;
  }});

  const sorted = Object.entries(agg).sort(
    (a, b) => (b[1].campus + b[1].ovcr + b[1].coe + b[1].dept)
            - (a[1].campus + a[1].ovcr + a[1].coe + a[1].dept)
  );

  // Build stack-friendly data
  const stackData = sorted.map(([dept, v]) => ({{
    department: dept, campus: v.campus, ovcr: v.ovcr, coe: v.coe, dept: v.dept
  }}));

  const depts = stackData.map(d => d.department);

  const stack = d3.stack().keys(CAT_KEYS);
  const series = stack(stackData);

  const margin = {{ top: 40, right: 20, bottom: 70, left: 80 }};
  const width = container.clientWidth - margin.left - margin.right;
  const height = 400;

  d3.select(container).select("svg").remove();

  const svg = d3.select(container).append("svg")
    .attr("width", width + margin.left + margin.right)
    .attr("height", height + margin.top + margin.bottom);

  const g = svg.append("g")
    .attr("transform", `translate(${{margin.left}},${{margin.top}})`);

  svg.append("text")
    .attr("class", "chart-title")
    .attr("x", (width + margin.left + margin.right) / 2)
    .attr("y", 26)
    .text("Indirect Cost Distribution by Department");

  const maxTotal = d3.max(stackData, d => d.campus + d.ovcr + d.coe + d.dept) || 1;

  const x = d3.scaleBand()
    .domain(depts)
    .range([0, width])
    .padding(0.25);

  const y = d3.scaleLinear()
    .domain([0, maxTotal * 1.1])
    .range([height, 0]);

  // X axis
  g.append("g")
    .attr("class", "axis")
    .attr("transform", `translate(0,${{height}})`)
    .call(d3.axisBottom(x))
    .selectAll("text")
    .attr("transform", "rotate(-40)")
    .style("text-anchor", "end");

  // Y axis
  g.append("g")
    .attr("class", "axis")
    .call(d3.axisLeft(y).ticks(6).tickFormat(formatDollars));

  // Y axis label
  g.append("text")
    .attr("class", "axis")
    .attr("transform", "rotate(-90)")
    .attr("y", -65)
    .attr("x", -height / 2)
    .attr("text-anchor", "middle")
    .style("font-size", "13px")
    .text("Indirect Cost ($)");

  // Stacked bars
  g.selectAll(".layer")
    .data(series)
    .enter().append("g")
    .attr("class", "layer")
    .attr("fill", d => CAT_COLORS[CAT_LABELS[d.key]])
    .selectAll("rect")
    .data(d => d.map(seg => ({{ ...seg, key: d.key }})))
    .enter().append("rect")
    .attr("x", d => x(d.data.department))
    .attr("width", x.bandwidth())
    .attr("y", height)
    .attr("height", 0)
    .attr("rx", 1)
    .on("mouseover", function(evt, d) {{
      d3.select(this).style("opacity", 0.8);
      const dept = d.data.department;
      const cat = CAT_LABELS[d.key];
      const val = d[1] - d[0];
      const total = d.data.campus + d.data.ovcr + d.data.coe + d.data.dept;
      showTooltip(evt,
        `<div class="tt-name">${{dept}}</div>` +
        `<div class="tt-dept" style="margin-bottom:4px">${{cat}}: ${{formatDollars(val)}}</div>` +
        `<div style="font-size:12px;color:#888">` +
        CAT_KEYS.map(k =>
          `<span style="color:${{CAT_COLORS[CAT_LABELS[k]]}}">\u25A0</span> ${{CAT_LABELS[k]}}: ${{formatDollars(d.data[k])}}`
        ).join("<br>") +
        `</div>` +
        `<div class="tt-amount" style="margin-top:4px;border-top:1px solid #eee;padding-top:4px">Total: ${{formatDollars(total)}}</div>`);
    }})
    .on("mousemove", positionTooltip)
    .on("mouseout", function() {{
      d3.select(this).style("opacity", 1);
      hideTooltip();
    }})
    .transition().duration(500).ease(d3.easeCubicOut)
    .attr("y", d => y(d[1]))
    .attr("height", d => y(d[0]) - y(d[1]));

  // Legend
  const legendDiv = document.getElementById("indirectLegend");
  legendDiv.innerHTML = "";
  CAT_KEYS.forEach(k => {{
    const item = document.createElement("div");
    item.className = "legend-item";
    const swatch = document.createElement("div");
    swatch.className = "legend-swatch";
    swatch.style.background = CAT_COLORS[CAT_LABELS[k]];
    item.appendChild(swatch);
    item.appendChild(document.createTextNode(CAT_LABELS[k]));
    legendDiv.appendChild(item);
  }});
}}

/* ---------- Quarter selector ---------- */
const quarterEl = document.getElementById("quarterSelect");
if (quarterEl) {{
  quarterEl.addEventListener("change", () => {{
    if (topInvState.svg) {{ topInvState.svg.remove(); topInvState.svg = null; }}
    renderTopInvestigators();
    renderDeptTotals();
    renderIndirectBreakdown();
  }});
}}

/* ---------- Initial render ---------- */
renderTopInvestigators();
renderDeptTotals();
renderIndirectBreakdown();
</script>
</body>
</html>"""

    outPath = outputDir / "dashboard.html"
    outPath.write_text(html, encoding="utf-8")
    return str(outPath)


def main():
    """CLI entry point. Reads combined Excel data and generates plots/dashboard."""
    parser = argparse.ArgumentParser(
        description="Generate summary plots from combined grant expenditure output."
    )
    parser.add_argument("input_file", help="Path to the combined Excel file (e.g. COE-Combined-Q1.xlsx)")
    parser.add_argument("--top", type=int, default=20,
                        help="Number of investigators in top-N chart (default: 20)")
    parser.add_argument("--format", dest="fmt", default="all",
                        choices=["png", "pdf", "html", "dashboard", "all"],
                        help="Output format: png, pdf, html, dashboard, or all (default: all)")
    parser.add_argument("-o", "--output-dir", default=".",
                        help="Directory for output files (default: current directory)")

    args = parser.parse_args()

    inputPath = Path(args.input_file)
    if not inputPath.exists():
        print(f"Error: Input file not found: {inputPath}", file=sys.stderr)
        sys.exit(1)

    outputDir = Path(args.output_dir)
    outputDir.mkdir(parents=True, exist_ok=True)

    print(f"Reading {inputPath}...")
    data = readCombinedData(inputPath)
    print(f"  {len(data)} investigators loaded")

    # Generate all three plots
    print("Generating plots...")
    mplTopInv, pltTopInv = plotTopInvestigators(data, topN=args.top)
    mplDeptTot, pltDeptTot = plotDeptTotals(data)
    mplIndirect, pltIndirect = plotIndirectBreakdown(data)

    saved = []

    # PNG output
    if args.fmt in ("png", "all"):
        for fig, name in [(mplTopInv, "topInvestigators"),
                          (mplDeptTot, "deptTotals"),
                          (mplIndirect, "indirectBreakdown")]:
            path = outputDir / f"{name}.png"
            fig.savefig(path, dpi=150, bbox_inches="tight")
            saved.append(str(path))

    # PDF output (all three in one multi-page PDF)
    if args.fmt in ("pdf", "all"):
        pdfPath = outputDir / "summary.pdf"
        with PdfPages(pdfPath) as pdf:
            for fig in [mplTopInv, mplDeptTot, mplIndirect]:
                pdf.savefig(fig, bbox_inches="tight")
        saved.append(str(pdfPath))

    # HTML output (interactive Plotly)
    if args.fmt in ("html", "all"):
        for pFig, name in [(pltTopInv, "topInvestigators"),
                           (pltDeptTot, "deptTotals"),
                           (pltIndirect, "indirectBreakdown")]:
            path = outputDir / f"{name}.html"
            pFig.write_html(str(path))
            saved.append(str(path))

    # Dashboard (single-page tabbed HTML)
    if args.fmt in ("html", "dashboard", "all"):
        dashPath = writeDashboard(data, outputDir)
        saved.append(dashPath)

    plt.close("all")

    print(f"Saved {len(saved)} files:")
    for f in saved:
        print(f"  {f}")


if __name__ == "__main__":
    main()
